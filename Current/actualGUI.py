# changes from lastest GUI to try to incorporate the logic for recollecting missed images, along with capturing images from OAK
# move RemoteCli to support folder

import os
import sys
import time
import pandas as pd
import openpyxl
import threading
import string
import shutil
import datetime
import paramiko
import socket
import cv2
import depthai as dai
import numpy as np
from datetime import date
import select
from dotenv import load_dotenv

from support.MachineMotion import *
sys.path.append("..")
from PyQt6 import QtWidgets
from PyQt6.QtCore import QTimer
from PyQt6.QtWidgets import *
from PyQt6.QtGui import QIntValidator

######## global vars fetched from .env file #########
load_dotenv()
# Length of the robot (between front and back sensors) measured in cm
ROBOT_LENGTH = int(os.environ['ROBOT_LENGTH'])
# width of the robot (distance between two front wheels) in cm
ROBOT_WIDTH = int(os.environ['ROBOT_WIDTH'])
# if you haven't changed your pi's name the default username is 'pi'
PI_USERNAME = os.environ['PI_USERNAME']
# if you haven't changed your pi's password, the default is 'raspberrypi'
PI_PASSWORD = os.environ['PI_PASSWORD']
# Distance between rows in mm
DISTANCE_TRAVELED = int(os.environ['DISTANCE_TRAVELED'])
# if BB moving backwards change the motors order to [3,2]
WHEEL_MOTORS = list(map(int, os.environ['WHEEL_MOTORS'].split(',')))
# number of ultrasonic sensors on the robot
NUMBER_OF_SENSORS = int(os.environ['NUMBER_OF_SENSORS'])
# json string describing the ultrasonic sensor pinout
ULTRASONIC_SENSOR_LISTS = json.dumps({
    # the trigger pins for sensors
    "trigger_pins": list(map(int, os.environ['TRIGGER_PINS'].split(','))),
    # the echo pins for sensors
    "echo_pins": list(map(int, os.environ['ECHO_PINS'].split(','))),
})
# reverse to move backwards or forwards
DIRECTIONS = os.environ['DIRECTIONS'].split(',')
# total distance between proximity sensors around camera plate
HOME_TO_END_SENSOR_DISTANCE = int(os.environ['HOME_TO_END_SENSOR_DISTANCE'])

######## Do not change these #########
# The server's hostname or IP address
PI_HOST = "192.168.7.3"
# The port used by the server
PI_PORT = 65432
# global flag used to stop execution
STOP_EXEC = False
# global flag used to mark completion
PROCESS_COMPLETE = False
# global flag to store the state
STATE = ""
# name of species sheet file
SPECIES_SHEET = "support\\SpeciesSheet.xlsx"
# name of images sheet file
IMAGES_SHEET = "support\\ImagesSheet.xlsx"
# path of the camera script
PATH = os.getcwd()+'\\support\\RemoteCli\\RemoteCli.exe'

############################### Ultrasonic sensors and laptop-pi communication functions ###############################

# Helper function for getting distance measurements from sensor


def get_distances(s, offsets):
    # Number of queries of sensor measurements
    dist_list = np.zeros((NUMBER_OF_SENSORS, NUMBER_OF_SENSORS))
    dist_raw = np.zeros((NUMBER_OF_SENSORS, NUMBER_OF_SENSORS))

    for k in range(0, NUMBER_OF_SENSORS):
        # Sending a request for data and receiving it
        s.sendall(bytes(ULTRASONIC_SENSOR_LISTS, encoding="utf-8"))

        # use timeout of 10 seconds to wait for sensor response
        ready = select.select([s], [], [], 10)
        if ready[0]:
            data = s.recv(4096)
            if "error" in str(data):
                return data

        else:
            print("Sensor error!")
            return "Sensor error!"
        time.sleep(0.25)

        # Parsing the measurements
        for i, item in enumerate(data.split()):
            dist_raw[k, i] = float(item)
            dist_list[k, i] = float(item) - offsets[i % 3]

    dist_list = np.median(dist_list, 0)
    return dist_list

# Helper function for computing orientation of robot


def find_orientation(distance):
    '''
    :param distance: np array. Average of N distances measured by the ultrasonic sensors. Array[0]=
    front right sensor, Array[1]= back right sensor.
    :param ROBOT_LENGTH: distance in cm from front right to back right sensor
    :return: angle of drift from straight trajectory
    '''
    print('distance[0]=', distance[0])
    print('distance[1]=', distance[1])
    theta = np.arctan2((distance[0]-distance[1]), ROBOT_LENGTH) * 180 / np.pi
    return theta

############################ End Ultrasonic sensors and laptop-pi communication functions #############################

########################################## Establish connection with pi #############################################


# INITIALIZING SERVER IN RPI
ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
ssh.connect(PI_HOST, username=PI_USERNAME, password=PI_PASSWORD)
ssh_stdin, ssh_stdout, ssh_stderr = ssh.exec_command(
    'python /home/{}/ultrasonic_calibration/RPI_ServerSensors.py &'.format(PI_USERNAME))
time.sleep(2)

# CONNECTING TO RPI SERVER

# Setting up the connection
s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
s.connect((PI_HOST, PI_PORT))


# LOADING OFFSETS
offsets = np.loadtxt('support\\SensorOffsets.csv', delimiter=',',
                     skiprows=1)  # delimiter added
print(offsets)


########################### OAK-D Funcs ###########################
def flushframes():
  for i in range(50):
    frame = queue.get()

def save_oak_image(timestamp):
  flushframes()
  filename = f"{STATE}_OAK_{timestamp}.jpg"
  frame = queue.get()
  imOut = frame.getCvFrame()
  cv2.imwrite(filename, imOut)

pipeline = dai.Pipeline()
camRgb = pipeline.create(dai.node.ColorCamera)
camRgb.setPreviewSize(3840, 2160)
camRgb.setBoardSocket(dai.CameraBoardSocket.RGB)
xoutRgb = pipeline.create(dai.node.XLinkOut)
camRgb.setResolution(dai.ColorCameraProperties.SensorResolution.THE_4_K)
camRgb.setColorOrder(dai.ColorCameraProperties.ColorOrder.RGB)
xoutRgb.setStreamName("rgb")
camRgb.preview.link(xoutRgb.input)
device = dai.Device(pipeline)
queue = device.getOutputQueue(name="rgb")
flushframes()

###################################################################


# main window class for selecting metadata


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.welcome_label = QLabel("WELCOME!")
        self.changes_label = QLabel(
            "Do you want to make changes to the species list?")
        self.state_label = QLabel('State')

        self.state_box = QComboBox()
        self.state_box.addItem('NC')
        self.state_box.addItem('TX')
        self.state_box.addItem('MD')

        self.yes_button = QPushButton('YES', self)
        self.yes_button.clicked.connect(self.opt_yes)
        self.no_button = QPushButton('NO', self)
        self.no_button.clicked.connect(self.opt_no)
        self.quit_button = QPushButton('Quit')
        self.quit_button.clicked.connect(QApplication.instance().quit)
        self.quit_button.resize(self.quit_button.sizeHint())
        self.quit_button.move(330, 260)

        self.grid = QGridLayout()
        self.grid.setSpacing(10)

        self.grid.addWidget(QLabel("            "), 0, 0)
        self.grid.addWidget(self.welcome_label, 1, 3)
        self.grid.addWidget(self.changes_label, 2, 2, 1, 3)
        self.grid.addWidget(self.state_label, 0, 4)
        self.grid.addWidget(self.state_box, 0, 5)
        self.grid.addWidget(self.yes_button, 3, 2)
        self.grid.addWidget(self.no_button, 3, 4)
        self.grid.addWidget(self.quit_button, 5, 5)

        self.setLayout(self.grid)

        self.setGeometry(100, 100, 600, 500)
        self.setWindowTitle('BenchBot')
        self.show()

    # take user to species page
    def opt_yes(self):
        global STATE
        STATE = self.state_box.currentText()
        page = SpeciesPage()
        main_window.addWidget(page)
        main_window.setCurrentIndex(main_window.currentIndex()+1)

    # take user to images page
    def opt_no(self):
        global STATE
        STATE = self.state_box.currentText()
        page = ImagesPage()
        main_window.addWidget(page)
        main_window.setCurrentIndex(main_window.currentIndex()+1)

# window class used to edit the species


class SpeciesPage(QWidget):
    def __init__(self):
        super().__init__()

        self.species_label = QLabel("Number of species to add/update?")
        self.total_species = QComboBox()

        self.species_list = [None] * 10
        self.row_list = [None] * 10

        for i in range(1, 10):
            self.total_species.addItem(str(i))
            self.species_list[i-1] = QLineEdit()
            self.row_list[i-1] = QLineEdit()

        self.total_species.currentIndexChanged.connect(self.update_list)
        self.earlier_count = 1

        self.species_label = QLabel("Name of Species")
        self.row_numbers = QLabel("Row Numbers")
        self.update_button = QPushButton('UPDATE', self)
        self.update_button.clicked.connect(self.check_cells)
        self.update_button.setEnabled(False)

        self.layout = QGridLayout()
        self.layout.addWidget(self.species_label, 0, 1, 1, 2)
        self.layout.addWidget(self.total_species, 0, 3)
        self.layout.addWidget(self.update_button, 11, 4)

        self.setLayout(self.layout)
        self.setGeometry(100, 100, 600, 500)
        self.setWindowTitle('BenchBot')

    # update species list
    def update_list(self):
        count = int(self.total_species.currentText())
        self.layout.addWidget(self.species_label, 1, 2)
        self.layout.addWidget(self.row_numbers, 1, 3)
        if(count > self.earlier_count):
            for current_count in range(self.earlier_count-1, count):
                self.layout.addWidget(
                    self.species_list[current_count], current_count+2, 2)
                self.layout.addWidget(
                    self.row_list[current_count], current_count+2, 3)
        if(count < self.earlier_count):
            for current_count in range(count, self.earlier_count):
                self.layout.removeWidget(self.species_list[current_count])
                self.layout.removeWidget(self.row_list[current_count])

        self.update_button.setEnabled(True)
        self.earlier_count = count

    # used to validate species
    def check_cells(self):
        # logic for checking contents of the fields before proceeding
        error_code = 0
        count = int(self.total_species.currentText())
        allowed_chars_1 = ['1', '2', '3', '4',
                           '5', '6', '7', '8', '9', '0', ',', '-']
        allowed_chars_2 = list(string.ascii_letters)
        allowed_chars_2.append(' ')

        for current_count in range(0, count):
            input_text_1 = self.row_list[current_count].text()
            input_text_2 = self.species_list[current_count].text()
            if input_text_1 == '' or input_text_2 == '':
                error_code = 1
            elif any(x not in allowed_chars_1 for x in input_text_1):
                error_code = 2
            elif any(x not in allowed_chars_2 for x in input_text_2):
                error_code = 3
            else:
                continue

        error_dialog = QMessageBox(self)
        error_dialog.setWindowTitle("Error")
        if error_code != 0:
            if error_code == 1:
                error_dialog.setText("Make sure all fields have entries")
            elif error_code == 2:
                error_dialog.setText(
                    "Enter row data in following format: '2' or '1-3' or '1,5,8' or '2-3,5'")
            else:
                error_dialog.setText(
                    "Make sure Species names only contain alphabets")
            error_dialog.setStandardButtons(QMessageBox.StandardButton.Close)
            if error_dialog.exec() == QMessageBox.StandardButton.Close:
                error_dialog.done(1)
        else:
            self.update_excel()

    # used to update SpeciesSheet.xlsx
    def update_excel(self):
        count = int(self.total_species.currentText())
        workbook = openpyxl.load_workbook(SPECIES_SHEET)
        sheet = workbook.active
        for current_count in range(0, count):
            sheet.cell(row=current_count+2,
                       column=1).value = self.species_list[current_count].text()
            sheet.cell(row=current_count+2,
                       column=2).value = self.row_list[current_count].text()
        for current_count in range(count, 9):
            sheet.cell(row=current_count+2, column=1).value = ''
            sheet.cell(row=current_count+2, column=2).value = ''
            sheet.cell(row=current_count+2, column=3).value = ''
        workbook.save(SPECIES_SHEET)
        os.system("python support\sheetupdateSpecies.py")
        # time.sleep(0.1)
        threading.Thread(target=backup_sheet).start()

        page = ImagesPage()
        main_window.addWidget(page)
        main_window.setCurrentIndex(main_window.currentIndex()+1)


# create a SpeciesSheet.xlsx backup


def backup_sheet():
    # create a copy of the sheet
    new_sheet = "support\\new.xlsx"
    shutil.copy(SPECIES_SHEET, new_sheet)
    
    # delete the images column
    workbook = openpyxl.load_workbook(new_sheet)
    sheet = workbook.active
    sheet.delete_cols(3, 1)
    workbook.save(new_sheet)

    # include date in file name
    today = datetime.datetime.today().strftime('%d-%b-%Y')
    os.rename(r'support\new.xlsx', r'support\SpeciesSheet_' + STATE + str(today) + '.xlsx')


# window class used to update the number of images per row


class ImagesPage(QWidget):
    def __init__(self):
        super().__init__()

        self.species_label = QLabel("Species")
        self.images_label = QLabel("Number of Images to Capture")
        self.next_button = QPushButton('NEXT', self)
        self.next_button.clicked.connect(self.check_content)
        self.only_int = QIntValidator(0, 8, self)

        self.layout = QGridLayout()
        self.layout.addWidget(QLabel("            "), 0, 0)
        self.layout.addWidget(self.species_label, 0, 1)
        self.layout.addWidget(self.images_label, 0, 2)
        self.layout.addWidget(self.next_button, 10, 3)

        species_df = pd.read_excel(SPECIES_SHEET)
        species_names = species_df[['SpeciesName']].values
        i = 0
        self.snaps = [None] * species_names.size
        for species_name in species_names:
            self.layout.addWidget(QLabel(species_name[0]), i+2, 1)
            self.snaps[i] = QLineEdit()
            self.snaps[i].setValidator(self.only_int)
            self.layout.addWidget(self.snaps[i], i+2, 2)
            i += 1

        self.setLayout(self.layout)
        self.setGeometry(100, 100, 600, 500)
        self.setWindowTitle('BenchBot')

    # validate the fields entered by user
    def check_content(self):
        error_code = 0
        for snap in range(0, len(self.snaps)):
            input_text = self.snaps[snap].text()
            if input_text == '':
                error_code = 1
            else:
                continue

        if error_code != 0:
            error_dialog = QMessageBox(self)
            error_dialog.setWindowTitle("Error")
            error_dialog.setText("Make sure all fields have entries")
            error_dialog.setStandardButtons(QMessageBox.StandardButton.Close)
            if error_dialog.exec() == QMessageBox.StandardButton.Close:
                error_dialog.done(1)
        else:
            self.confirm_message()

    # verify the user is ready to start and then run acquisition
    def confirm_message(self):
        confirm_dialog = QMessageBox(self)
        confirm_dialog.setWindowTitle("Confirm")
        confirm_dialog.setText(
            "Are you sure you want to begin the image acquisition process?")
        confirm_dialog.setStandardButtons(
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        confirm_dialog.setIcon(QMessageBox.Icon.Question)
        button = confirm_dialog.exec()

        if button == QMessageBox.StandardButton.Yes:
            workbook = openpyxl.load_workbook(SPECIES_SHEET)
            sheet = workbook.active
            for snap in range(0, len(self.snaps)):
                sheet.cell(
                    row=snap+2, column=3).value = self.snaps[snap].text()
            workbook.save(SPECIES_SHEET)
            os.system("python support\sheetupdatePictures.py")
            confirm_dialog.done(1)
            page = AcquisitionPage()
            main_window.addWidget(page)
            main_window.setCurrentIndex(main_window.currentIndex()+1)
        else:
            confirm_dialog.done(1)


# window class that handles the acquisition process and updates the user on what's happening


class AcquisitionPage(QWidget):
    def __init__(self):
        super().__init__()

        # Initialize the machine motion object
        self.mm = MachineMotion(DEFAULT_IP)
        self.mm.releaseEstop()
        self.mm.resetSystem()
        self.camera_motor = 1

        self.acquisition_label = QLabel("Image Acquisition Process Started")
        self.time_label = QLabel("     Time Elapsed: 0 hrs 0 mins")
        self.cancel_button = QPushButton('CANCEL', self)
        self.cancel_button.clicked.connect(self.stop_thread)

        self.layout = QGridLayout()
        self.layout.addWidget(QLabel("            "), 0, 0)
        self.layout.addWidget(self.acquisition_label, 0, 1)
        self.layout.addWidget(self.time_label, 1, 1)
        self.layout.addWidget(self.cancel_button, 4, 1)
        self.layout.addWidget(QLabel("            "), 3, 2)

        self.setLayout(self.layout)
        self.setGeometry(100, 100, 600, 500)
        self.setWindowTitle('BenchBot')

        self.start_time = time.time()
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_label)
        self.timer.start(60000)  # every 1 min

        threading.Thread(target=self.acquisition_process).start()

    def configure_machine_motion(self):
        # config machine motion
        self.mm.configAxis(
            self.camera_motor, MICRO_STEPS.ustep_8, MECH_GAIN.ballscrew_10mm_turn)
        self.mm.configAxisDirection(
            self.camera_motor, DIRECTION.POSITIVE)

        for axis in WHEEL_MOTORS:
            self.mm.configAxis(
                axis, MICRO_STEPS.ustep_8, MECH_GAIN.enclosed_timing_belt_mm_turn)
            self.mm.configAxisDirection(axis, DIRECTIONS[axis-2])

        self.mm.emitAcceleration(50)
        self.mm.emitSpeed(80)

    def generate_pots_list(self):
        images_df = pd.read_excel(IMAGES_SHEET)
        image_counts = images_df[['ImagesCount']].values
        rows = []
        for num in image_counts:
            rows.append(num[0])
        return rows

    def create_directory(self):
        # Create new directory where today's images will be saved.
        directory_name = f'{STATE}_{date.today()}'
        if directory_name not in os.listdir():
            os.mkdir(f'{os.getcwd()}\{directory_name}')
        os.chdir(f'{os.getcwd()}\{directory_name}')
        os.mkdir(f'{os.getcwd()}\OAK')
        os.mkdir(f'{os.getcwd()}\SONY')

    def correct_path(self):
        corrected_distance = get_distances(s, offsets)
        if "error" in corrected_distance:
            print(corrected_distance)
            return

        # Getting angle
        ang = find_orientation(corrected_distance)
        print('debug_angle=', ang)

        # Calculate how much distance motor needs to move to align platform
        d_correction_mm = 2*np.pi*ROBOT_WIDTH*(abs(ang)/360)*10
        print('debug_d_correction_mm=', d_correction_mm)

        # Create if statement to indicate which motor moves
        if ang > 0.5:
            self.mm.moveRelative(
                WHEEL_MOTORS[0], d_correction_mm)
            self.mm.waitForMotionCompletion()
        elif ang < -0.5:
            self.mm.moveRelative(
                WHEEL_MOTORS[1], d_correction_mm)
            self.mm.waitForMotionCompletion()
    
    def capture_image(self):
        t = str(int(time.time()))
        os.startfile(PATH)
        save_oak_image(t)
        time.sleep(8)
        threading.Thread(target=self.file_rename(t)).start()
    
    def file_rename(self, timestamp):
        time.sleep(4)
        # t = str(int(time.time()))
        for file_name in os.listdir('.'):
            if file_name.startswith(STATE+'X'):
                if file_name.endswith('.JPG'):
                    new_name = f"{STATE}_{timestamp}.JPG"
                elif file_name.endswith('.ARW'):
                    new_name = f"{STATE}_{timestamp}.ARW"
                os.rename(f"{file_name}", f"{new_name}")
            else:
                continue

    def process_finished(self):
        global PROCESS_COMPLETE
        PROCESS_COMPLETE = True
        self.mm.moveToHome(self.camera_motor)
        self.mm.waitForMotionCompletion()
        self.mm.triggerEstop()
        self.acquisition_label.setText('Acquisition Process Complete')
        current_time = time.time()
        elapsed_time = int(current_time-self.start_time)
        elapsed_hr = int(elapsed_time/3600)
        elapsed_min = int(elapsed_time/60) - 60*elapsed_hr
        self.time_label.setText(
            'Total time taken: '+str(elapsed_hr) + ' hrs '+str(elapsed_min) + ' mins')

    def stop_thread(self):
        stopping = threading.Thread(target=self.stop)
        stopping.start()

    def stop(self):
        global STOP_EXEC
        STOP_EXEC = True
        self.mm.moveToHome(self.camera_motor)
        self.mm.waitForMotionCompletion()
        self.mm.triggerEstop()
        if not PROCESS_COMPLETE:
            self.acquisition_label.setText('Acquisition Process Disrupted')
            self.time_label.setText('     Close the Application')

    def update_label(self):
        if PROCESS_COMPLETE or STOP_EXEC:
            self.timer.stop()
        else:
            current_time = time.time()
            elapsed_time = int(current_time-self.start_time)
            elapsed_hr = int(elapsed_time/3600)
            elapsed_min = int(elapsed_time/60) - 60*elapsed_hr
            self.time_label.setText(
                '     Time Elapsed: '+str(elapsed_hr) + ' hrs '+str(elapsed_min) + ' mins')

    def round(self, pots, distance):
        for i in range(1, pots):
            if STOP_EXEC:
                break
            # Trigger capture of image
            self.capture_image()
            # Move camera plate to next point
            self.mm.moveRelative(self.camera_motor, distance)
            self.mm.waitForMotionCompletion()
        # Trigger image capture at last point            
        self.capture_image()

    def repeat_round(self, expected_count):
        filelist = [name for name in os.listdir('.') if os.path.isfile(name)]
        actual_count = len(filelist)
        if ((expected_count*3)>actual_count):
            # scrub directory of the files
            for file in filelist:
                os.remove(file)            
            return True
        else:
            return False

    def move_files(self):
        sony_folder = os.getcwd()+'\\OAK'
        oak_folder = os.getcwd()+'\\SONY'
        for file in os.listdir('.'):
            if file.endswith('.JPG') or file.endswith('.ARW'):
                shutil.move(file, sony_folder)
            elif file.endswith('.jpg'):
                shutil.move(file, oak_folder)
            else:
                continue

    # moves benchbot and captures images
    def acquisition_process(self):
        self.configure_machine_motion()
        total_rows = self.generate_pots_list()
        self.create_directory()
        direction = True
        # expected_images = 0

        for pots in total_rows:
            self.correct_path()
            if pots == 0 or pots == 1:
                if STOP_EXEC:
                    break
                self.mm.moveRelativeCombined(WHEEL_MOTORS, [DISTANCE_TRAVELED, DISTANCE_TRAVELED])
                self.mm.waitForMotionCompletion()
            else:
                # expected_images += pots
                # total distance between home and end sensor
                total_distance = int(HOME_TO_END_SENSOR_DISTANCE/(pots-1))

                if not direction:
                    total_distance *= -1
                self.round(pots, total_distance)

                if self.repeat_round(pots):
                    total_distance *= -1
                    direction = not direction
                    self.round(pots, total_distance)
                
                self.move_files()

                direction = not direction
                self.mm.moveRelativeCombined(WHEEL_MOTORS, [DISTANCE_TRAVELED, DISTANCE_TRAVELED])
                self.mm.waitForMotionCompletion()

        if STOP_EXEC:
            self.stop()
        else:
            self.process_finished()



if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = QtWidgets.QStackedWidget()
    window = MainWindow()
    main_window.addWidget(window)
    main_window.setFixedHeight(400)
    main_window.setFixedWidth(500)
    main_window.show()
    sys.exit(app.exec())